from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json

file_path = r"C:\Users\rlsug\Downloads\VSN_Satsang 2 -Allocation_21-03-2026.xlsx"
wb = load_workbook(file_path)

print("=" * 100)
print("WORKSHEETS IN FILE:")
print("=" * 100)

for sheet_index, sheet_name in enumerate(wb.sheetnames):
    print(f"\n{'='*100}")
    print(f"TAB {sheet_index + 1}: {sheet_name}")
    print(f"{'='*100}")
    
    ws = wb[sheet_name]
    print(f"Max Row: {ws.max_row}, Max Col: {ws.max_column}\n")
    
    # Print all data with cell references
    print("DATA:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        row_data = []
        for cell in row:
            value = cell.value
            # Get cell color/fill
            fill_color = cell.fill.fgColor.rgb if cell.fill else None
            row_data.append({
                'value': value,
                'fill': fill_color
            })
        
        # Print row
        row_str = " | ".join([str(item['value']) if item['value'] else '' for item in row_data])
        print(f"Row {row_idx}: {row_str}")
    
    print("\n" + "-" * 100 + "\n")

print("\n✓ Excel file analyzed successfully!")
